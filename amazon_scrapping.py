#selenium imports
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException,NoSuchElementException
#for chrome web driver setting
from webdriver_manager.chrome import ChromeDriverManager
#for solving Amazon CAPTCHAs
from amazoncaptcha import AmazonCaptcha
#importing bs4
from bs4 import BeautifulSoup
#for joining Links
from urllib.parse import urljoin
#time module
import time
#pandas for data evaluation
import pandas as pd
#for some excel operations
from openpyxl import load_workbook


#setting chrome options
chrome_options=Options()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--disable-gpu') #Disable gpu acceleration (optional)
chrome_options.add_argument("--disable-dev-shm-usage")  # Overcome limited resource problems
#Setting up webdriver
driver=webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=chrome_options)


#declaring list to hold data
product_list=[]
#defining some elements
max_attempts=3        
attempt=0                   #attempt number initialization
captcha_solved=False        #bool to track captcha solving
timeout=8                   #duration for timeout 
product_count=0             #product counter

search_element="Laptops"    #category to be searched
num_pages=10                #number of pages to access from website

#----------------------------------------------------------------------------------------------------------------------------------------
#---------------------------------------------------FUNCTIONS----------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------------------------------------------

#function to solve captchas
def solve_captcha():
    try:
        captcha_img = WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class = 'a-row a-text-center']//img")))
        captcha_link=captcha_img.get_attribute('src')
        captcha=AmazonCaptcha.fromlink(captcha_link)
        captcha_val=AmazonCaptcha.solve(captcha)
        time.sleep(1)
        input_box=driver.find_element(By.ID,"captchacharacters")
        input_box.send_keys(captcha_val)
        button=driver.find_element(By.CLASS_NAME,"a-button-text")
        button.click()
        return True                #captcha solved
    except NoSuchElementException:
        return False               #captcha not found

#function to get titles of the products
def get_title(pd,h2_tag):
    title=h2_tag.find('span',class_='a-size-medium a-color-base a-text-normal')
    if not title:
        title=pd.find('span',class_='a-color-base')
        
    return title.text.strip() if title and title.text else "No title available"

#function to get link of the products    
def get_link(h2_tag,base_url):
    relative_url=h2_tag.a['href']
    return urljoin(base_url,relative_url) if relative_url else " No Link Available"

#function to get rating of the products
def get_rating(pd):
    ratings=pd.find('span',class_='a-icon-alt')
    rating=ratings.text.strip() if ratings else "Ratings not available"
    
    return rating.replace('out of 5 stars','') 

#function to get prices (original and featured(if available)) of the products
def get_prices(pd):
    price_span = pd.find('span', class_='a-price')
    if price_span:
        f_price = price_span.find('span', class_='a-offscreen')  #main displaying price (storing in f_price for now)
        if f_price:
            f_price_text = f_price.text.strip()
        else:
            f_price_text="No Featured Offers"
        
        p_span=pd.find('span', class_='a-price a-text-price') #if the original price tag is found beside the displaying price (featured price exist)
        if p_span:
            o_price=p_span.find('span', class_='a-offscreen')
            if o_price:
                o_price_text=o_price.text.strip() #getting formatted text in string 
            else:
                o_price_text="No Price Available"
        else:
            o_price=f_price  #if original price tag is not found beside displaying price then the DISPLAYING price is the ORIGINAL price(no featured prices)
            if o_price:
                o_price_text=o_price.get_text(strip=True)
                f_price_text="No Featured Offers"
            else:
                o_price_text = "No original price available"
    else:
        f_price_text = "No featured Offers"        # if nothing is found in the main price fields then look for other fields where the price would be located
        op_span=pd.find('div',{'data-cy':'secondary-offer-recipe'})
        if op_span:
            o_price=op_span.find('span',class_='a-color-base')
            o_price_text=o_price.text.strip()
        else:
            o_price_text = "No original price available"
    
    if o_price_text.startswith('$') and f_price_text.startswith('$'):
        org_price=float(o_price_text.replace("$",'').replace(',','').strip())
        ft_price=float(f_price_text.replace("$",'').replace(',','').strip())
        
        if org_price<ft_price:
            o_price_text=f_price_text
            f_price_text="No featured Offers" 
    
    return o_price_text,f_price_text
    
#----------------------------------------------------------------------------------------------------------------------------------------
#---------------------------------------------------MAIN WORKING-------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------------------------------------------
try:
    #retry attempts in case of any error
    while attempt<max_attempts and not captcha_solved:  
        
        #defining url
        BASE_URL='https://www.amazon.com'
        driver.get(BASE_URL)
        time.sleep(1)
        
        try:
            #solving Amazon Captcha if it appears
            captcha_solved=solve_captcha()
            print("Captcha detected and solved!")
            #waiting for the search box to appear on the main page
            WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, "twotabsearchtextbox"))
            )
            search_box = driver.find_element(By.ID, "twotabsearchtextbox")
            search_box.clear()
            break  
        except (TimeoutException, NoSuchElementException):
            attempt += 1                    
            print(f"Attempt {attempt}/{max_attempts} failed, retrying...")   
            driver.quit()
    #in case all attempts failed        
    if attempt>max_attempts:
        print("Failed to solve captcha or load the main page after maximum attempts.")
    else:
        #sending the category to be searched in the searched box
        search_box.send_keys(search_element)
        search_button=driver.find_element(By.ID,"nav-search-submit-button").click()
    
    #page wise scrapping
    for page in range(num_pages):
        try:
            #additional time in case of slow internet connection to load the contents of the page successfully
            time.sleep(1.5)
            #converting from selenium to BeautifulSoup for easy parsing
            html_data=driver.page_source
            soup=BeautifulSoup(html_data,'lxml') 
            
            #accessing the products cards list
            products=soup.find_all('div',{'data-cy':'asin-faceout-container'})
                
            #looping through each product to get data
            for product in products:
                
                #getting h2 tag first
                h2_tag=product.find('h2')
                #getting title of each product
                product_title=get_title(product,h2_tag)
                #getting link to each product
                product_link=get_link(h2_tag,BASE_URL)
                #Getting rating of each product
                product_rating=get_rating(product)
                #getting price for both featured and non-featured products
                original_price,featured_price=get_prices(product)
                
                product_count+=1
                #appending data in dictionary
                product_data={
                    "Title": product_title,
                    "Original Price": original_price,
                    "Featured Price": featured_price,
                    "Rating":product_rating,
                    "Link": product_link
                }
                #appending dictionary to list
                product_list.append(product_data)
            
            
            # Check if the "Next" button is present
            next_button = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "s-pagination-next"))
            )
            next_button.click()
            print(f"Page # {page+1} successfully scraped!")
        
        except TimeoutException:
            print(f"Skipping the page # {page + 1} as 'NEXT' Button is not found")
            continue    #skipping and moving to the next page
        except Exception as e:
            print(f"Unexpected error occur on page # {page + 1} : {e}")
            break
        
except Exception as e:
    print(e)
finally:
        driver.quit()


#printing data from dictionary in console
print("Total Data : ",product_count,'\n\n')
for p in product_list:
    print(f'Title : {p["Title"]} \nOriginal Price: {p["Original Price"]} \nFeatured Price: {p["Featured Price"]} \nRating: {p["Rating"]} \nLink : {p["Link"]}\n\n')

#Excel operations

if product_list:
    #building data frame with pandas for further editing
    dataframe=pd.DataFrame(product_list)

    #defining file and sheet names for excel
    file_name=f'amazon_{search_element}_data.xlsx'
    sheet_name=f'{search_element}_Data'


    try:
        #writing data to excel file
        dataframe.to_excel(file_name,sheet_name=sheet_name,index=False)
        print(f"Data saved to Excel File : {file_name}")
        
        #making link clickable in excel sheet
        wb=load_workbook(file_name)                                                 #accessing the workbook
        ws=wb[sheet_name]                                                           #accessing the specific excel sheet
        sheet_col=dataframe.columns.get_loc('Link')+1
        
        for row_no in range(2,ws.max_row+1):                                        #assuming that first row is occupied by headers
            cell=ws.cell(row=row_no,column=sheet_col)                               #defining the cell position for iteration
            if cell.value:
                cell.hyperlink=cell.value
                cell.style='Hyperlink'
                
        wb.save(file_name)                                                          #saving the edited file
    except Exception as e:
        print(f"Couldn't write data to excel file \nError : {e}")
else:
    print("No data scraped..Skipping Excel File operations")
